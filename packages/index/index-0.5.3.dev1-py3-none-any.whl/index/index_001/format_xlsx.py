#!/usr/bin/env python
# coding=utf-8
# Stan 2024-11-01

import datetime

from openpyxl import load_workbook

from ..chunk import chunk
from ..timer import Timer
from .funcs import get_shid_name


def main_yield(filename, db, options={}, **kargs):
    with Timer(f"[ {__name__} ] load_workbook", db.verbose) as t:
        book = load_workbook(filename, data_only=True)

    sheet_names = book.get_sheet_names()
    sheet_list  = options.get('sheets', sheet_names)
    chunk_rows  = options.get('chunk_rows', 5000)
    row_mode    = options.get('row_mode', 1)
    cells_mode  = options.get('cells_mode', 0)

    processed = []

    for name in sheet_list:         # 1-based integer or string
        # 1-based integer and string
        shid, shname = get_shid_name(sheet_names, name)
        if shid is None:
            db.push_file_record('warning',
                message = f"Wrong sheet name: {name}"
            )
            continue

        if shid in processed:
            db.push_file_record('info',
                message = f"Sheet already processed: {name}"
            )
            continue

        processed.append(shid)

        sh = book[shname]               # string
#       sh = book.worksheets[shid-1]    # 0-based

        if db.verbose:
            print(f"Processing: # {shid} ({sh.title}) / max_row: {sh.max_row}, max_column: {sh.max_column}")

        for ki, chunk_i in enumerate(chunk(sh.iter_rows(), chunk_rows)):
            records = []

            for kj, row in enumerate(chunk_i):
                record = {}
                if row_mode:
                    record['row'] = get_row_values(row)

                if cells_mode:
                    record['_cells'] = get_cells(row)

                record = {k: v for k, v in record.items() if v}
                if record:
                    idx = ki * chunk_rows + kj
                    _r = idx + 1

                    record = dict(**record, _r=_r)
                    records.append(record)

            yield records, {
                '_shid': shid,
#               '_shname': sh.title
            }
            records = []        # release memory

    book.close()


# Plain mode
def get_row_values(row):
    values = [parse_cell(cell) for cell in row]

    if any(x is not None for x in values):
        return values


def parse_cell(cell):
    if cell.data_type == 'n':   # TYPE_NUMERIC, TYPE_NULL
        return cell.value

    if cell.data_type == 's':   # TYPE_STRING
        return cell.value.strip()

    if cell.data_type == 'str': # TYPE_FORMULA_CACHE_STRING
        if cell.value is None:
            return ''

        return cell.value

    if cell.data_type == 'd':   # ext: cell.is_date
        if isinstance(cell.value, datetime.time):
            return f"TIME({cell.value})"

        if isinstance(cell.value, datetime.timedelta):
            return f"TIMEDELTA({cell.value})"

        return cell.value

    if cell.data_type == 'b':   # TYPE_BOOL
        return cell.value

    if cell.data_type == 'e':   # TYPE_ERROR
        return f"ERROR({cell.value})"

    if cell.data_type == 'f':   # TYPE_FORMULA
        return f"FORMULA({cell.value})"

#   if cell.data_type == 'inlineStr':   # TYPE_INLINE

    return cell.value


# Attribute pattern
def get_cells(row):
    values = []
    for col_i, cell in enumerate(row, 1):
        value = parse_cell_ext(cell, col_i)

        note_dict = get_note(cell)
        if note_dict:
            if value:   # dict
                value['_n'] = note_dict

            else:
                value = dict(c=col_i, _n=note_dict)

        values.append(value)
    values = [x for x in values if x is not None]

    return values


def parse_cell_ext(cell, col_i):
    if cell.data_type == 'n':   # TYPE_NUMERIC, TYPE_NULL
        val = cell.value
        if val is None:
            return

        return {
            'c': col_i,
            'v': val,
            't': 2
        }

    if cell.data_type == 's':   # TYPE_STRING
        val = cell.value.strip()
        return {
            'c': col_i,
            'v': val,
            't': 1
        }

    if cell.data_type == 'str': # TYPE_FORMULA_CACHE_STRING
        val = cell.value
        if val is None:
            return {
                'c': col_i,
                'v': '',
                't': 1
            }

        return {
            'c': col_i,
            'v': val,
            't': -1         # trick for rare type
        }

    if cell.data_type == 'd':   # ext: cell.is_date
        val = cell.value
        if isinstance(val, datetime.time):
            val = f"TIME({val})"
            return {
                'c': col_i,
                'v': val,
                't': -3         # temporary solution
            }

        if isinstance(cell.value, datetime.timedelta):
            val = f"TIMEDELTA({val})"
            return {
                'c': col_i,
                'v': val,
                't': -4         # trick for rare type
            }

        return {
            'c': col_i,
            'v': val,
            't': 3
        }

    if cell.data_type == 'b':   # TYPE_BOOL
        val = cell.value
        return {
            'c': col_i,
            'v': val,
            't': 4
        }

    if cell.data_type == 'e':   # TYPE_ERROR
        val = cell.value
        return {
            'c': col_i,
            'v': val,
            't': 5
        }

    if cell.data_type == 'f':   # TYPE_FORMULA
        val = cell.value
        return {
            'c': col_i,
            'v': val,
            't': 6
        }

#   if cell.data_type == 'inlineStr':   # TYPE_INLINE

    val = cell.value
    return {
        'c': col_i,
        'v': val,
        't': cell.data_type     # trick for rare type
    }


def get_note(cell):
    if cell.comment:
        return dict(
            author = cell.comment.author,
            text = cell.comment.text
        )
