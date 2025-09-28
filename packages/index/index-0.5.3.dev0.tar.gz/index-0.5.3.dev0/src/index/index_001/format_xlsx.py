#!/usr/bin/env python
# coding=utf-8
# Stan 2024-11-01

import datetime

from openpyxl import load_workbook

from ..chunk import chunk
from ..timer import Timer
from .funcs import get_shid_name


def main_yield(filename, db, options={}, **kargs):
    with Timer(f"[ {__name__} ] load_workbook", db.verbose) as t:
        book = load_workbook(filename, read_only=True, data_only=True)

    sheet_names = book.get_sheet_names()
    sheet_list  = options.get('sheets', sheet_names)
    chunk_rows  = options.get('chunk_rows', 5000)

    for name in sheet_list:         # 1-based integer or string
        # 1-based integer and string
        shid, shname = get_shid_name(sheet_names, name)
        if shid is None:
            db.push_task_record('warning', f"Wrong sheed name/id: {name}")

        sh = book[shname]               # string
#       sh = book.worksheets[shid-1]    # 0-based

        if db.verbose:
            print(f"Processing: # {shid} ({sh.title}) / max_row: {sh.max_row}, max_column: {sh.max_column}")

        for ki, chunk_i in enumerate(chunk(sh.iter_rows(), chunk_rows)):
            records = []

            for kj, row in enumerate(chunk_i):
                row_values = get_row_values(row)
                if row_values:
                    idx = ki * chunk_rows + kj
                    _r = idx + 1

                    record = dict(row=row_values, _r=_r)
                    records.append(record)

            yield records, {
                '_shid': shid,
#               '_shname': sh.title
            }
            records = []        # release memory

    book.close()


def get_row_values(row):
    values = [parse_cell(i) for i in row]
    if any(x is not None for x in values):
        return values

    return []


def parse_cell(cell):
    if cell.data_type == 'n':   # Numeric, empty cell
        return cell.value

    if cell.data_type == 's':   # str
        return cell.value.strip()

    if cell.data_type == 'str': # empty string =""
        if cell.value is None:
            return ''
        else:
            return cell.value

    if cell.data_type == 'd':   # and cell.is_date
        if isinstance(cell.value, datetime.time):
            return f"TIME({cell.value})"
        else:
            return cell.value

    if cell.data_type == 'b':   # bool
        return cell.value

    if cell.data_type == 'e':   # Error
        return f"ERROR({cell.value})"

    if cell.data_type == 'f':   # DataTableFormula, ArrayFormula
        return f"FORMULA({cell.value})"

    return cell.value
