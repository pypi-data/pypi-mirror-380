#!/usr/bin/env python
# coding=utf-8
# Stan 2024-10-29

import os
from pprint import pformat

from openpyxl import load_workbook

from .timer import Timer


def get_definition(filename, suit_name, **kargs):
    verbose = kargs.get('verbose')
    debug   = kargs.get('debug')

    res = get_description(filename, **kargs)
    if debug:
        print("get_description:")
        print(pformat(res, 2, 160))
        print("=======")

    nrows       = res['nrows']
    rows_values = res['rows_values']
    key_map     = res['key_map']
    suit_names   = res['suit_names']

    rows_values, keys = get_set_filtered(rows_values, key_map, suit_names, suit_name)

    return nrows, rows_values, keys


def get_description(filename, **kargs):
    verbose = kargs.get('verbose')
    debug   = kargs.get('debug')

    with Timer(f"[ {__name__} ] load_workbook '{filename}'", verbose) as t:
        book = load_workbook(filename, read_only=True, data_only=True)

    if debug:
        print(f"sheetname list of description file: {book.sheetnames}\n")

    sh = book.worksheets[0]

    nrows       = 0
    rows_idx    = []
    rows_values = []
    set_idx     = []
    suit_names   = []
    key_map     = []

    for i, row in enumerate(sh.rows, 1):
        if i == 1:
            for j, cell in enumerate(row):
                if cell.value == 'c':
                    nrows += 1
                    rows_idx.append(j)

                elif cell.value:
                    if cell.value in suit_names:
                        raise Exception(f"Key duplicates: {cell.value}")

                    set_idx.append(j)
                    suit_names.append(cell.value)

        else:
            rows_values.append( tuple([i.value for i in (map(row.__getitem__, rows_idx))]) )
            key_map.append([i.value for i in (map(row.__getitem__, set_idx))])

    return {
        'nrows':       nrows,
#       'rows_idx':    rows_idx,
        'rows_values': rows_values,
#       'set_idx':     set_idx,
        'suit_names':   suit_names,
        'key_map':     key_map
    }


def get_set(key_map, suit_names, suit_name):
    idx = suit_names.index(suit_name)

    return [i[idx] for i in key_map]


def filter_columns(rows_values, keys_raw):
    index_list = [idx for idx, val in enumerate(keys_raw) if val != None]

    rows = map(rows_values.__getitem__, index_list)
    keys = map(keys_raw.__getitem__, index_list)

    return list(rows), list(keys)


def get_set_filtered(rows_values, key_map, suit_names, suit_name):
    keys_raw = get_set(key_map, suit_names, suit_name)

    return filter_columns(rows_values, keys_raw)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Check table definition")

    parser.add_argument('filename',
                        help="specify a file",
                        metavar="file.xlsx")

    parser.add_argument('--set',
                        required=True,
                        help="specify the name of the set",
                        metavar="key1")

    parser.add_argument('-v', '--verbose',
                        action='store_true',
                        help="verbose mode")

    parser.add_argument('--debug',
                        action='store_true',
                        help="debug mode")

    args = parser.parse_args()

    if args.debug:
        print("Command line arguments:")
        for key, value in vars(args).items():
            print(f"- {key:16}: {value}")
        print()

    if not os.path.isfile(args.filename):
        raise FileNotFoundError(f"File not found: '{args.filename}'")

    res = get_definition(
        args.filename,
        args.set,
        verbose = args.verbose,
        debug   = args.debug
    )

    if args.verbose:
        print("Res:", pformat(res, 2, 160), end="\n\n")


if __name__ == '__main__':
    main()
