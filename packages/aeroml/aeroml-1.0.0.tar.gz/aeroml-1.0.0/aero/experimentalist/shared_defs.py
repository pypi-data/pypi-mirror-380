import math
import re
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import List, Dict, Any, TypedDict, Annotated, Tuple
from langchain_core.messages import BaseMessage
from langgraph.graph import add_messages
import pandas as pd 
from langgraph.config import get_stream_writer
import inspect
from openpyxl import load_workbook 
@dataclass
class Evidence:
    snippet: str
    source: str
    score: float


@dataclass
class PropertyHit:
    name: str
    evidence: List[Evidence]
    
    @property
    def confidence(self) -> float:
        """Calculate confidence based on evidence."""
        if not self.evidence:
            return 0.0
        
        # Calculate base confidence using independent signals
        prod = 1.0
        for ev in self.evidence:
            prod *= (1.0 - max(0.0, min(1.0, ev.score)))
        base_confidence = 1.0 - prod
        
        # Apply evidence count bonus with diminishing returns
        evidence_bonus = min(0.05 * math.log(len(self.evidence) + 1), 0.15)
        
        final_confidence = min(1.0, base_confidence + evidence_bonus)
        return round(final_confidence, 3)
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary format."""
        return {
            "name": self.name,
            "confidence": self.confidence,
            "evidence": [asdict(ev) for ev in self.evidence],
        }


class BaseState(TypedDict):
    """Base state for all workflows."""
    messages: Annotated[List[BaseMessage], add_messages]
    original_prompt: str  # Pure user query without uploaded data
    uploaded_data: List[str]  # Uploaded file contents as separate field
    current_step: str
    errors: List[str]


  
class ExperimentSuggestionState(BaseState):
    """State object for the experiment suggestion workflow."""
    # Input data
    experimental_results: Dict[str, Any]      # Raw experimental data
    findings_analysis: Dict[str, Any]         # Analysis of current findings
    research_context: Dict[str, Any]          # Context about the research domain
    
    # Processing state
    analysis_completed: bool                  # Whether initial analysis is done
    experiment_categories: List[str]          # Types of experiments identified
    experiment_papers: List[Dict[str, Any]]   # Papers retrieved for experimental guidance
    experiment_search_query: str              # Query used for paper search
    experiment_search_iteration: int          # Current search iteration count
    experiment_validation_results: Dict[str, Any]  # Results from experiment validation (not paper validation)
    experiment_paper_validation_decision: str # Decision from validation (continue/search_new/search_backup)
    experiment_validation_decision: str       # Overall validation decision (PASS/FAIL)
    experiment_iterations: List[Dict[str, Any]]  # History of experiment iterations
    research_direction: Dict[str, Any]        # Research direction analysis
    validated_experiment_papers: List[Dict[str, Any]]  # Validated papers for suggestions
    distilled_methodologies: Dict[str, Any]         # Distilled methodology content from papers
    current_experiment_iteration: int        # Current iteration of experiment suggestion
    iteration_from_state: int                 # Iteration number from state
    analysis_iterations: List[Dict[str, Any]] # Track analysis validation iterations (history)
    direction_iterations: List[Dict[str, Any]] # Track research direction validation iterations (history)
    # Issue tracking for iterative improvement
    past_fixed_issues: List[str]              # Issues that were resolved in previous iterations
    past_unresolved_issues: List[str]         # Issues that persist across iterations
    most_recent_generation_issues: List[str]  # Issues from the most recent experiment generation
    cumulative_validation_feedback: List[Dict[str, Any]]  # Historical validation feedback
    
    # 🆕 PAST MISTAKES TRACKING FOR ITERATIVE LEARNING
    past_experiment_mistakes: List[Dict[str, Any]]  # Historical validation failures for LLM learning
    
    # Output
    experiment_suggestions: str                # Comprehensive experiment suggestions
    experiment_summary: Dict[str, Any]         # Summary of experiment generation
    next_node: str                            # Next node to route to in workflow
    literature_context: str                    # Extracted literature context for experiments
    suggestion_source: str                     # Source of the experiment suggestions
    prioritized_experiments: List[Dict[str, Any]]  # Ranked experiment list
    implementation_roadmap: Dict[str, Any]    # Step-by-step implementation plan
    final_outputs: Dict[str, str]             # Final formatted outputs  

    # Dependencies needed by workflow nodes
    client: Any                                # OpenAI client
    model: str                                 # Model name
    arxiv_processor: Any                       # ArxivPaperProcessor instance





# ==================================================================================
# STREAMWRITER HELPER FUNCTION
# ==================================================================================

def _write_stream(message: str, key: str = "status"):
    """Helper function to write to StreamWriter if available."""
    try:
        # Use LangGraph's get_stream_writer() without parameters (proper way)
        writer = get_stream_writer()
        writer({key: message})
    except Exception:
        # Fallback: try to get stream from config (for testing compatibility)
        try:
            # This fallback is for test compatibility only
           
            frame = inspect.currentframe()
            while frame:
                if 'config' in frame.f_locals and frame.f_locals['config']:
                    config = frame.f_locals['config']
                    stream = config.get("configurable", {}).get("stream")
                    if stream and hasattr(stream, 'write'):
                        stream.write(message)
                        return
                frame = frame.f_back
        except Exception:
            pass
        # Final fallback: silently fail
        pass



# ===== End moved from shared_constants.py =====

def _clean_text_for_utf8(text):
    """Clean text to ensure UTF-8 compatibility by removing surrogate characters."""
    if not isinstance(text, str):
        return str(text)
    
    # Remove surrogate characters that cause UTF-8 encoding issues
    # Remove surrogate pairs (Unicode range U+D800-U+DFFF)
    text = re.sub(r'[\ud800-\udfff]', '', text)
    
    # Replace other problematic Unicode characters with safe alternatives
    text = text.encode('utf-8', errors='ignore').decode('utf-8')
    
    # Clean up any remaining control characters except newlines and tabs
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    
    return text




def create_custom_ranking_prompt(prompt_type: str = "default") -> str:
    """Create a custom ranking prompt based on prompt type."""
    
    if prompt_type == "experimental":
        return """
            You are an expert experimental methodology researcher.  
            Your task: Estimate how relevant this paper is to **experimental research needs** using ONLY the paper’s title and summary (abstract).  

            OUTPUT FORMAT (STRICT):
            - Return EXACTLY one floating-point number with ONE decimal (e.g., 8.7).  
            - No words, no JSON, no units, no symbols, no explanation.  
            - Single line only (no leading/trailing spaces or extra lines).  

            SCORING CRITERIA (use inference from title/summary):  
            - methodology_relevance (40%): Does the summary explicitly mention experimental methodology, benchmarks, protocols, or evaluation setups?  
            - experimental_evidence (30%): Does it mention results, experiments, performance comparisons, or ablation studies?  
            - implementation_guidance (20%): Does it provide or strongly imply practical details like datasets, code availability, reproducibility, or implementation notes?  
            - research_alignment (10%): Does it align with the given research direction and questions?  

            COMPUTE:  
            - Let m,e,i,r ∈ [0,1], estimated from the title/summary.  
            - score = round((0.40*m + 0.30*e + 0.20*i + 0.10*r) * 10, 1).  
            - If the title/summary clearly lacks experimental content (all four < 0.15), output **1.0**.  
            - Clip final result to [1.0, 10.0].  

            PRIORITIZATION:  
            - Favor papers with explicit mention of **empirical studies, benchmarks, datasets, or evaluation frameworks**.  
            - Penalize papers that are purely theoretical, conceptual, or survey-style with no experimental grounding.  
            Research context:
            \"\"\"{query}\"\"\"

            Paper title:
            \"\"\"{title}\"\"\"

            Paper summary:
            \"\"\"{content}\"\"\"
        """.strip()
    
    elif prompt_type == "model_suggestion":
        return """
            You are an expert ML model selection researcher. Score how relevant this paper is to model selection and architecture research on a 1–10 scale.

            OUTPUT FORMAT (STRICT):
            - Return EXACTLY one floating-point number with ONE decimal (e.g., 8.7).
            - No words, no JSON, no units, no symbols, no explanation.
            - Single line only (no leading/trailing spaces or extra lines).

            MODEL FOCUS SCORING - assign four subscores in [0,1]:
            - architecture_relevance (40%): discusses relevant model architectures, neural network designs, or ML approaches
            - performance_evidence (30%): provides performance benchmarks, comparisons, or evaluation results
            - implementation_details (20%): includes implementation specifics, hyperparameters, training procedures, or code
            - task_alignment (10%): addresses similar tasks, domains, or application requirements

            Compute:
            - Let a,p,i,t ∈ [0,1].
            - score = round((0.40*a + 0.30*p + 0.20*i + 0.10*t) * 10, 1).
            - If clearly unrelated to models/architectures (all four < 0.15), output 1.0.
            - Clip to [1.0, 10.0].
            - Prioritize papers with concrete model architectures and performance data.

            Research context:
            \"\"\"{query}\"\"\"

            Paper title:
            \"\"\"{title}\"\"\"

            Paper summary:
            \"\"\"{content}\"\"\"
        """.strip()
                    
    else:  # default prompt
        return None  # Use the original prompt in arxiv_paper_utils.py



def _load_text_file_safely(file_path: str) -> Tuple[List[str], List[str]]:
    """Attempt to load file content as UTF-8 and gracefully skip binary data.

    Returns a tuple of (text_snippets, warnings)."""

    warnings: List[str] = []
    if not file_path:
        return [], warnings

    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    suffix = path.suffix.lower()

    if suffix in {".xlsx", ".xls", ".xlsm"}:
       
        try:
            excel_data = pd.read_excel(path, sheet_name=None)
        except Exception as exc:  # pragma: no cover - pandas-specific errors
            warnings.append(
                f"⚠️ Failed to parse Excel file '{path.name}': {exc}; skipping attachment content."
            )
            return [], warnings

        text_snippets: List[str] = []
        workbook = None
        openpyxl_failed = False

        for sheet_name, sheet_df in excel_data.items():
            usable_df = None
            if sheet_df is not None:
                trimmed_df = sheet_df.dropna(how="all")
                if not trimmed_df.empty:
                    trimmed_df = trimmed_df.loc[:, trimmed_df.notna().any()]
                if not trimmed_df.empty:
                    usable_df = trimmed_df

            if usable_df is None or usable_df.empty:
                try:
                    alt_df = pd.read_excel(path, sheet_name=sheet_name, header=None)
                    alt_df = alt_df.dropna(how="all")
                    if not alt_df.empty:
                        alt_df = alt_df.loc[:, alt_df.notna().any()]
                except Exception:  # pragma: no cover - pandas-specific errors
                    alt_df = None
                if alt_df is not None and not alt_df.empty:
                    usable_df = alt_df
                    warnings.append(
                        f"Excel sheet '{sheet_name}' parsed without headers due to sparse data."
                    )

            if usable_df is None or usable_df.empty:
                if workbook is None and not openpyxl_failed:
                    try:
                         # type: ignore

                        workbook = load_workbook(path, read_only=True, data_only=True)
                    except Exception:  # pragma: no cover - openpyxl specific errors
                        workbook = None
                        openpyxl_failed = True

                if workbook is not None and sheet_name in workbook.sheetnames:
                    rows: List[str] = []
                    for row in workbook[sheet_name].iter_rows(values_only=True):
                        if not row or all(
                            (cell is None)
                            or (isinstance(cell, str) and not cell.strip())
                            for cell in row
                        ):
                            continue
                        row_values = ["" if cell is None else str(cell) for cell in row]
                        rows.append(",".join(row_values))
                        if len(rows) >= 200:
                            break
                    if rows:
                        raw_text = "\n".join(rows)
                        cleaned_rows = _clean_text_for_utf8(raw_text)
                        snippet = f"Sheet: {sheet_name}\n{cleaned_rows}"
                        if len(snippet) > 10000:
                            snippet = snippet[:10000] + "\n... (truncated)"
                            warnings.append(
                                f"ℹ️ Excel sheet '{sheet_name}' truncated to 10k characters."
                            )
                        warnings.append(
                            f"ℹ️ Excel sheet '{sheet_name}' extracted via openpyxl fallback."
                        )
                        text_snippets.append(snippet)
                        continue

            if usable_df is None or usable_df.empty:
                continue

            preview_df = usable_df.head(200)
            if len(usable_df) > len(preview_df):
                warnings.append(
                    f"Excel sheet '{sheet_name}' large; attached first {len(preview_df)} rows only."
                )
            preview = preview_df.to_csv(index=False)
            cleaned = _clean_text_for_utf8(preview)
            snippet = f"Sheet: {sheet_name}\n{cleaned}"
            if len(snippet) > 10000:
                snippet = snippet[:10000] + "\n... (truncated)"
                warnings.append(
                    f"Excel sheet '{sheet_name}' truncated to 10k characters."
                )
            text_snippets.append(snippet)

        if not text_snippets:
            warnings.append(
                f"Excel file '{path.name}' contained no tabular data to attach."
            )

        return text_snippets, warnings

    if suffix in {".csv", ".tsv"}:
    
        sep = "\t" if suffix == ".tsv" else ","
        try:
            df = pd.read_csv(path, sep=sep)
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(path, sep=sep, encoding="latin-1")
                warnings.append(
                    "CSV decoding failed with UTF-8; loaded using latin-1 encoding."
                )
            except Exception as exc:
                warnings.append(
                    f"Failed to parse CSV/TSV file '{path.name}': {exc}; falling back to raw text."
                )
                df = None
        except Exception as exc:  # pragma: no cover - pandas-specific errors
            warnings.append(
                f"Failed to parse CSV/TSV file '{path.name}': {exc}; falling back to raw text."
            )
            df = None

        if df is not None:
            preview_df = df.head(200)
            if len(df) > len(preview_df):
                warnings.append(
                    f"CSV/TSV file '{path.name}' large; attached first {len(preview_df)} rows only."
                )
            preview_text = preview_df.to_csv(index=False)
            cleaned = _clean_text_for_utf8(preview_text)
            
            return [cleaned], warnings

    # Fallback for text files
    try:
        with path.open("r", encoding="utf-8") as handle:
            return [_clean_text_for_utf8(handle.read())], warnings
    except UnicodeDecodeError:
        try:
            size = path.stat().st_size
        except OSError:
            size = None
        size_note = f" ({size} bytes)" if size is not None else ""
        warnings.append(
            f"Detected non-text input file '{path.name}'{size_note}; skipping attachment content."
        )
    except FileNotFoundError:
        raise FileNotFoundError(f"Input file not found: {file_path}")
    except Exception as exc:
        raise Exception(f"Error reading input file: {exc}")

    return [], warnings
