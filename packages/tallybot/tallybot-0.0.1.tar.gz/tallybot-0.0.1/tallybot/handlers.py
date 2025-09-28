"""Functions to handle various email attachment types and data conversions."""

import datetime
import io
import json
import os
import re
import string
import zipfile

import pdftotext  # type: ignore[import-not-found]
import openpyxl  # typpe: ignore


def get_content(msg):
    """Get content from email msg type.

    Return content as as tuple of (data, attachment)
    """
    data = {}
    attachment = b""
    supported = ["csv", "zip", "xlsx", "pdf"]
    for i in msg.walk():
        cur_type = i.get_content_subtype()
        if cur_type in supported or cur_type == "plain":
            content = i.get_content()
            if "plain" == cur_type:
                data = get_json(content)
            elif cur_type in ["csv", "zip", "xlsx", "pdf"]:
                attachment = content
    return data, attachment


# pylint: disable=R0903
class ExcelGenerator:
    """Generates Excel."""

    def __init__(self, *r_structs):
        """From Report structure will generate Excel binary."""
        self.__r_structs = r_structs
        self.__write_excel_sheets()

    def __write_excel_sheets(self):
        """Convert r_structs to Excel rows format."""
        self.__sheets = []
        for r_struct in self.__r_structs:
            rows = []
            rows += r_struct.header
            start = r_struct.start
            for i, j in enumerate(r_struct.attrs):
                rows.append(
                    (f"{string.ascii_uppercase[i]}{str(start)}", j.capitalize())
                )
            for a, item in enumerate(r_struct.items):
                for i, j in enumerate(r_struct.attrs):
                    val = getattr(item, j)
                    if j in r_struct.callbacks:
                        try:
                            val = r_struct.callbacks[j](val)
                        except AttributeError:
                            raise RuntimeWarning(
                                f"Transaction {item} misses partner"
                            ) from None
                    rows.append(
                        (
                            f"{string.ascii_uppercase[i]}{str(start + a + 1)}",
                            val,
                        )
                    )
            self.__sheets.append((rows, r_struct.title))

    def binary(self):
        """Prepare binary stream object.

        Requires a rows (cell data to be written in tuple (cell, data))
        That is prepared into Excel file and returned as binary file-
        object.
        """
        excel = openpyxl.Workbook()
        first = True
        for sheet in self.__sheets:
            if first:
                ws_sheet = excel.active
                ws_sheet.title = sheet[1]
                first = False
            else:
                ws_sheet = excel.create_sheet(title=sheet[1])
            for cell in sheet[0]:
                ws_sheet[cell[0]].value = cell[1]
        tmp = io.BytesIO()
        excel.save(tmp)
        tmp.seek(0)
        excel_bytes = tmp.read()
        tmp.close()
        return excel_bytes


def get_excel(content):
    """From bytes content returns Openpyxl object."""
    bin_read = io.BytesIO()
    bin_read.write(content)
    bin_read.seek(0)
    excel = openpyxl.load_workbook(bin_read)
    bin_read.close()
    return excel


def get_json(text):
    """From plain text 'key:value format' builds json string."""
    content = text.split("\n")
    json_string = "["
    data = "{"
    previous = False
    header = True
    for i, j in enumerate(content):
        col = j.find(":")
        if col == -1:
            continue
        val = j[col + 1 :].strip()
        key = j[0:col].strip().lower()
        if key == previous:
            json_string += data[:-1] + "},"
            data = "{"
        data += f'"{key}": "{val}",'
        if i == len(content) - 1:
            json_string += data[:-1] + "}"
            data = ""
        if header:
            previous = key
            header = False
    if data and data[-1] == ",":
        json_string += data[:-1] + "}"
    json_string += "]"
    try:
        json_string = json.loads(json_string)
    except json.JSONDecodeError:
        raise RuntimeWarning(f"Corrupt json format {json_string}")
    return json_string


def get_zip(content):
    """Return a list of all files in zip archive."""
    fn = io.BytesIO()
    fn.write(content)
    fn.seek(0)
    with zipfile.ZipFile(fn) as archive:
        result = []
        for i in archive.namelist():
            with archive.open(i) as temp:
                result.append(temp.read())
    fn.close()
    return result


def get_pdf(content):
    """Return pdf object."""
    fn = io.BytesIO(content)
    # pylint: disable=I1101
    pdf = pdftotext.PDF(fn, physical=True)
    fn.close()
    return pdf


def get_my_invoice(text):
    """Return invoice data from text."""
    patterns = {
        "invoice_no": r"INVOICE #\s*(\w+)",
        "date": r"INVOICE DATE\s*(.+)",
        "amount": r"TOTAL AMOUNT\s*EUR\s*(\d*,*\d+\.\d{2})",
        "partner": r"Bill to:\s*(.+)",
    }
    for key, p in patterns.items():
        res = re.search(p, text)
        if res:
            patterns[key] = res.group(1)
        else:
            patterns[key] = ""
    patterns["no"] = re.search(r"\d+", patterns["invoice_no"]).group()
    patterns["amount"] = patterns["amount"].replace(",", "")
    patterns["date"] = datetime.datetime.strptime(
        patterns["date"].replace(" ", ""), "%d%b%Y"
    )
    return patterns


# pylint: disable=R0912,R0915
def get_invoice(text):
    """Return invoice data from text."""
    invoice_date = reference = comment = value = issuer = partner = ""
    for i, j in enumerate(text):
        if j == "F":
            char = text[i : i + 5]
            if char == "From:":
                issuer = text[i + 6 : i + 100].strip().split("\n")[0]
        elif j == "B":
            if text[i : i + 8] == "Bill to:":
                partner = text[i + 9 : i + 50].strip()
        elif j == "I":
            char = text[i : i + 9]
            if char == "INVOICE #":
                word_start = False
                word_end = False
                word = ""
                j = 0
                while not all((word_start, word_end)) and j < 50:
                    j += 1
                    char = text[i + 9 + j].strip()
                    if char:
                        word += char
                        word_start = True
                    if not char and word_start:
                        word_end = True
                reference = word.strip()
        elif j == "D":
            char = text[i - 5 : i + 8]
            if char == "     DATE    ":
                invoice_date = text[i + 4 : i + 42].strip()
            char = text[i : i + 18]
            if char == "DESCRIPTION / MEMO":
                a = 18
                while True:
                    char = text[i + a]
                    if comment == "AMOUNT":
                        comment = ""
                        break
                    if len(comment) > 150:
                        raise RuntimeWarning(
                            "Could not find properly comment field"
                        )
                    comment += char
                    comment = comment.strip()
                    a += 1
                comment = text[i + a : i + a + 150]
                comment = comment.strip()
                comment = comment.split("   ")[0]
        elif j == "T":
            char = text[i - 5 : i + 17]
            if char == " " * 5 + "TOTAL AMOUNT" + " " * 5:
                value = text[i + 12 : i + 39].strip()[1:]
    if not all((invoice_date, reference, comment, value, issuer, partner)):
        err_msg = "Could not fetch all required data from invoice\n"
        err_msg += f"Invoice date: {invoice_date}\n"
        err_msg += f"Reference: {reference}\n"
        err_msg += f"Comment: {comment}\n"
        err_msg += f"Value: {value}\n"
        err_msg += f"Issuer: {issuer}\n"
        err_msg += f"Partner: {partner}\n\n{text}"
        raise RuntimeWarning(err_msg)
    invoice_date = datetime.datetime.strptime(invoice_date, "%b %d, %Y")
    value = value.replace(",", "")
    return (
        invoice_date,
        reference,
        comment,
        value,
        issuer,
        partner,
    )


def save_file(full_path, binary):
    """Save binary file as per full_path."""
    if os.path.exists(full_path):
        raise RuntimeWarning(f"File already exists in {full_path}")
    with open(full_path, mode="wb") as fname:
        fname.write(binary)
