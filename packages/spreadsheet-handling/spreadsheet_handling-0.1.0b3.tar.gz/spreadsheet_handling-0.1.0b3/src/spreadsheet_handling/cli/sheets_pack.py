#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pack (Phase 1+2a/2b):
- Mehrere JSON-Quellen -> ein Workbook (XLSX) oder ein CSV-Ordner.
- FK-Helper-Spalten (Engine.apply_fks) und Validierung (Engine.validate).
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Sequence

import pandas as pd

from spreadsheet_handling.core.fk import assert_no_parentheses_in_columns
from spreadsheet_handling.engine.orchestrator import Engine
from spreadsheet_handling.logging_utils import setup_logging, get_logger
from spreadsheet_handling.pipeline.pipeline import run_pipeline, build_steps_from_config


log = get_logger("pack")

DEFAULTS: Dict[str, Any] = {
    "levels": 3,
    "backend": "xlsx",  # xlsx|csv
    "id_field": "id",  # ID-Feld in Zielblättern
    "label_field": "name",  # menschenlesbares Label
    "helper_prefix": "_",
    "detect_fk": True,
}


# ---------- Utilities ----------


def _ensure_multiindex(df: pd.DataFrame, levels: int) -> pd.DataFrame:
    """
    Stellt sicher, dass df.columns ein MultiIndex mit 'levels' Ebenen ist.
    Wenn flache Spaltennamen vorliegen, wird die erste Ebene belegt und der Rest mit "" aufgefüllt.
    """
    if isinstance(df.columns, pd.MultiIndex):
        if df.columns.nlevels < levels:
            new_tuples = []
            for tpl in df.columns.to_list():
                if not isinstance(tpl, tuple):
                    tpl = (tpl,)
                fill = ("",) * (levels - len(tpl))
                new_tuples.append(tuple(tpl) + fill)
            df.columns = pd.MultiIndex.from_tuples(new_tuples)
        return df

    cols = list(df.columns)
    tuples = [(c,) + ("",) * (levels - 1) for c in cols]
    df.columns = pd.MultiIndex.from_tuples(tuples)
    return df


def _read_json_records(path: Path) -> List[Dict[str, Any]]:
    """
    Liest eine JSON-Datei. Erlaubt als Root entweder Liste[Objekt] oder einzelnes Objekt.
    """
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return [data]
    raise ValueError(f"Unsupported JSON root in {path}: {type(data)}")


# ---------- Config Handling ----------


def _load_config(args: argparse.Namespace) -> Dict[str, Any]:
    """
    Lädt YAML-Config (wenn --config gesetzt) oder baut eine ad-hoc-Konfiguration
    aus <input-dir> und -o/--output.

    Schema (v1):
    {
      workbook: str,
      defaults: {levels:int, backend:str, ...},
      sheets: [ {name:str, json:str} | {json:str(dir)} ]
    }
    """
    if args.config:
        try:
            import yaml  # type: ignore
        except ImportError as exc:  # pragma: no cover
            raise SystemExit("PyYAML ist nicht installiert. Bitte `pip install pyyaml`.") from exc

        with open(args.config, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f) or {}

        # Defaults mergen
        dfl = DEFAULTS.copy()
        dfl.update((cfg.get("defaults") or {}))
        cfg["defaults"] = dfl

        if not cfg.get("workbook"):
            if not args.output:
                raise SystemExit("`workbook` fehlt (YAML) und -o/--output ist nicht gesetzt.")
            cfg["workbook"] = args.output
        return cfg

    # No-YAML-Modus
    if not args.input or not args.output:
        raise SystemExit("Ohne --config brauchst du <json_dir> und -o <workbook>.")

    json_dir = Path(args.input)
    if not json_dir.is_dir():
        raise SystemExit(f"{json_dir} ist kein Verzeichnis")

    backend = args.backend or DEFAULTS["backend"]
    cfg = {
        "workbook": args.output,
        "defaults": {
            "levels": args.levels or DEFAULTS["levels"],
            "backend": backend,
        },
        "sheets": [],
    }

    # Jede *.json Datei wird ein Blatt
    for p in sorted(json_dir.glob("*.json")):
        cfg["sheets"].append({"name": p.stem, "json": str(p)})

    if not cfg["sheets"]:
        raise SystemExit(f"Keine *.json Dateien in {json_dir} gefunden.")
    return cfg


# ---------- Frame-Lader ----------


def _load_frames_from_jsons(cfg: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    """Konvertiert alle in cfg angegebenen JSON-Quellen in DataFrames (mit MultiIndex-Spalten)."""
    defaults = cfg.get("defaults", {})
    levels = int(defaults.get("levels", DEFAULTS["levels"]))

    frames: Dict[str, pd.DataFrame] = {}
    for sheet_cfg in cfg.get("sheets", []):
        src = Path(sheet_cfg["json"])  # Datei ODER Verzeichnis
        if src.is_dir():
            # Erzeuge je Datei ein Blatt (Name=Stem)
            for p in sorted(src.glob("*.json")):
                records = _read_json_records(p)
                df = pd.DataFrame(records)
                frames[p.stem] = _ensure_multiindex(df, levels)
        else:
            name = sheet_cfg.get("name") or src.stem
            records = _read_json_records(src)
            df = pd.DataFrame(records)
            frames[name] = _ensure_multiindex(df, levels)

    return frames


# ---------- Writer ----------



def _write_xlsx(workbook_path: Path, frames: Dict[str, pd.DataFrame]) -> None:
    """
    Excel: MultiIndex-Spalten robust schreiben, indem wir die Spalten
    vor dem Schreiben auf die 1. Ebene flatten (eine Headerzeile).
    Danach: AutoFilter + dezente Header-Färbung auf jeder Tabelle.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    workbook_path = workbook_path.with_suffix(".xlsx")

    # 1) schreiben (wie bisher)
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as xw:
        for sheet, df in frames.items():
            df_out = df.copy()
            if isinstance(df_out.columns, pd.MultiIndex):
                # nur die erste Ebene verwenden (Level 0)
                df_out.columns = [t[0] for t in df_out.columns.to_list()]
            df_out.to_excel(xw, sheet_name=sheet, index=False)

    # 2) nachbearbeiten: AutoFilter + Header-Styling
    wb = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)

    for ws in wb.worksheets:
        # AutoFilter über den genutzten Bereich (inkl. Headerzeile)
        # ws.dimensions z. B. "A1:D100"
        ws.auto_filter.ref = ws.dimensions

        # Header-Zeile = 1 (wir schreiben mit header=0)
        max_col = ws.max_column or 0
        if max_col > 0:
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font

    wb.save(workbook_path)
    print(f"[pack] XLSX geschrieben: {workbook_path}")

def _write_csv_folder(out_dir: Path, frames: Dict[str, pd.DataFrame]) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    for sheet, df in frames.items():
        df_out = df.copy()
        if isinstance(df_out.columns, pd.MultiIndex):
            # Nur Level-0 in die CSV-Headerzeile schreiben
            df_out.columns = [t[0] for t in df_out.columns.to_list()]
        df_out.to_csv(out_dir / f"{sheet}.csv", index=False, encoding="utf-8")
    print(f"[pack] CSV-Ordner geschrieben: {out_dir}")


def _write_workbook(cfg: Dict[str, Any], frames: Dict[str, pd.DataFrame]) -> None:
    out = cfg.get("workbook")
    if not out:
        raise SystemExit("Output-Pfad `workbook` fehlt.")
    out_path = Path(out)

    backend = (cfg.get("defaults", {}).get("backend") or DEFAULTS["backend"]).lower()
    if backend == "xlsx":
        _write_xlsx(out_path, frames)
    elif backend == "csv":
        # out kann Datei- oder Ordnername sein -> immer zu Ordner normalisieren
        if out_path.suffix:
            out_path = out_path.parent / out_path.stem
        _write_csv_folder(out_path, frames)
    else:
        raise SystemExit(f"Unbekannter Backend-Typ: {backend}")


# ---------- Orchestrator (Library Entry) ----------


def run_pack(
    cfg: Dict[str, Any],
    *,
    mode_missing_fk: str | None = None,
    mode_duplicate_ids: str | None = None,
) -> None:
    """
    Orchestriert: Frames laden -> validieren -> FK-Helpers -> schreiben.

    Die Validierungsmodi werden priorisiert:
      1) explizite Funktionsargumente
      2) cfg['defaults']['validate'] (Schlüssel: 'missing_fk', 'duplicate_ids')
      3) Fallback 'warn'
    """
    defaults: Dict[str, Any] = cfg.get("defaults", {}) or {}
    log.debug("run_pack defaults=%s", defaults)

    # 1) Frames laden
    frames: Dict[str, pd.DataFrame] = _load_frames_from_jsons(cfg)
    log.info("loaded %d sheet(s): %s", len(frames), list(frames.keys()))

    # 2) Header-Guards (keine Klammern)
    for sheet_name, df in frames.items():
        assert_no_parentheses_in_columns(df, sheet_name)

    # 3) Engine initialisieren
    engine = Engine(defaults)

    # 3a) Validierungsmodi bestimmen (CLI/Funktionsargs > defaults.validate > 'warn')
    vcfg = defaults.get("validate") or {}
    mmode = mode_missing_fk if mode_missing_fk is not None else vcfg.get("missing_fk", "warn")
    dmode = (
        mode_duplicate_ids if mode_duplicate_ids is not None else vcfg.get("duplicate_ids", "warn")
    )

    # 3b) Validieren (wir loggen den Report in DEBUG, raisen je nach Modus in Engine.validate)
    report = engine.validate(frames, mode_missing_fk=mmode, mode_duplicate_ids=dmode)
    log.debug("validate report=%s", report)

    # 3c) FK-Helper-Spalten anwenden (nur wenn detect_fk=True)
    frames = engine.apply_fks(frames)

    # 4) Schreiben (xlsx/csv/…)
    _write_workbook(cfg, frames)


# ---------- CLI ----------


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Pack JSON -> Workbook (Multi-Sheet)")
    p.add_argument("input", nargs="?", help="JSON-Verzeichnis (no-YAML-Modus)")
    p.add_argument("-o", "--output", help="Workbook-Pfad (xlsx) oder Ordner (csv)")
    p.add_argument("--config", help="YAML-Konfiguration")
    p.add_argument("--levels", type=int, default=None, help="Header-Levels (default 3)")
    p.add_argument("--backend", choices=["xlsx", "csv"], help="xlsx (default) oder csv")
    p.add_argument(
        "--missing-fk",
        choices=["ignore", "warn", "fail"],
        help="Validation mode for missing foreign keys (default: warn).",
    )
    p.add_argument(
        "--duplicate-ids",
        choices=["ignore", "warn", "fail"],
        help="Validation mode for duplicate IDs (default: warn).",
    )
    p.add_argument(
        "--fail-on-missing-fk", action="store_true", help="Shortcut for --missing-fk=fail"
    )
    p.add_argument(
        "--fail-on-duplicate-ids", action="store_true", help="Shortcut for --duplicate-ids=fail"
    )
    p.add_argument(
        "--log-level",
        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
        help="Logger-Level (default WARNING)",
    )
    return p


def main(argv: Sequence[str] | None = None) -> int:
    ap = build_arg_parser()
    args = ap.parse_args(argv)
    setup_logging(args.log_level)

    cfg = _load_config(args)

    # Validierungsmodi ermitteln (Priorität: CLI > defaults.validate > 'warn')
    defaults = cfg.get("defaults", {})
    vcfg = defaults.get("validate", {}) or {}
    mode_missing = (
        args.missing_fk
        or ("fail" if args.fail_on_missing_fk else None)
        or vcfg.get("missing_fk")
        or "warn"
    )
    mode_dup = (
        args.duplicate_ids
        or ("fail" if args.fail_on_duplicate_ids else None)
        or vcfg.get("duplicate_ids")
        or "warn"
    )

    run_pack(
        cfg,
        mode_missing_fk=mode_missing,
        mode_duplicate_ids=mode_dup,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
