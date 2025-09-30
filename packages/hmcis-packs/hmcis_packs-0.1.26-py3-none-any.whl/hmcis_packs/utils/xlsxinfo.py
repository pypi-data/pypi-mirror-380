import warnings
from pathlib import Path
from typing import Dict, Any

from openpyxl import load_workbook


class SimpleExcelAnalyzer:
    """
    Упрощенная версия анализатора Excel файлов
    Полагается на read_only=True для решения проблем блокировки
    """

    def __init__(self, file_path: str):
        """
        Args:
            file_path (str): Путь к Excel файлу
        """
        self.file_path = Path(file_path).resolve()
        self.workbook = None
        self._open_workbook()

    def _open_workbook(self):
        """Открывает Excel файл"""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Файл не найден: {self.file_path}")

        try:
            # Подавляем предупреждения openpyxl
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

                # read_only=True должен решить проблемы с блокировкой
                self.workbook = load_workbook(
                    filename=str(self.file_path),
                    read_only=True,  # Ключевой параметр!
                    data_only=True  # Читаем значения, не формулы
                )

            print(f"✅ Файл открыт: {self.file_path.name}")
            print(f"📊 Листов: {len(self.workbook.sheetnames)}")

        except PermissionError as e:
            # Если все-таки не получилось, выводим понятную ошибку
            raise Exception(
                f"Файл заблокирован и не может быть прочитан: {self.file_path}\n"
                f"Попробуйте:\n"
                f"1. Закрыть файл в Excel\n"
                f"2. Проверить права доступа\n"
                f"3. Скопировать файл в другое место\n"
                f"Ошибка: {e}"
            )
        except Exception as e:
            raise Exception(f"Ошибка при открытии файла: {e}")

    def get_basic_info(self) -> Dict[str, Any]:
        """Получает базовую информацию о файле"""
        file_stats = self.file_path.stat()

        info = {
            'file_name': self.file_path.name,
            'file_size_mb': round(file_stats.st_size / (1024 * 1024), 2),
            'sheets_count': len(self.workbook.sheetnames),
            'sheet_names': self.workbook.sheetnames,
        }

        # Размеры листов
        dimensions = {}
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                dimensions[sheet_name] = (0, 0)
            else:
                dimensions[sheet_name] = (ws.max_row, ws.max_column)

        info['sheet_dimensions'] = dimensions
        return info

    def get_detailed_info(self) -> Dict[str, Any]:
        """Получает подробную информацию"""
        basic = self.get_basic_info()

        # Добавляем метаданные
        props = self.workbook.properties
        if props.creator:
            basic['author'] = props.creator
        if props.title:
            basic['title'] = props.title
        if props.modified:
            basic['modified'] = props.modified.strftime('%Y-%m-%d %H:%M:%S')

        # Проверяем на макросы (простая проверка по расширению)
        basic['has_macros'] = self.file_path.suffix.lower() == '.xlsm'

        return basic

    def print_summary(self):
        """Выводит краткую сводку"""
        print("=" * 50)
        print("📊 АНАЛИЗ EXCEL ФАЙЛА")
        print("=" * 50)

        info = self.get_detailed_info()

        print(f"\n📁 Файл: {info['file_name']}")
        print(f"💾 Размер: {info['file_size_mb']} MB")
        print(f"📄 Листов: {info['sheets_count']}")

        if info.get('author'):
            print(f"👤 Автор: {info['author']}")
        if info.get('modified'):
            print(f"📅 Изменен: {info['modified']}")

        print(f"\n📋 Листы:")
        for i, (name, (rows, cols)) in enumerate(info['sheet_dimensions'].items(), 1):
            print(f"  {i}. '{name}' - {rows} строк × {cols} столбцов")

    def close(self):
        """Закрывает файл"""
        if self.workbook:
            self.workbook.close()
            print("📁 Файл закрыт")

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def __del__(self):
        self.close()


# Функция для быстрого анализа
def quick_analyze(file_path: str) -> Dict[str, Any]:
    """
    Быстрый анализ Excel файла без создания объекта

    Args:
        file_path (str): Путь к файлу

    Returns:
        Dict с информацией о файле
    """
    try:
        with SimpleExcelAnalyzer(file_path) as analyzer:
            return analyzer.get_detailed_info()
    except Exception as e:
        return {'error': str(e)}


if __name__ == "__main__":
    print("🧪 ТЕСТИРОВАНИЕ ПРОСТОГО ПОДХОДА")
    print("=" * 50)

    # Пример использования
    file_path = r"V:\Accounting\Work\Мерц\2025\3 квартал 2025\Август 2025 года\Отчетность\!Начисление МСФО_август 2025.xlsx"  # Замените на реальный файл

    try:
        # Способ 1: Контекстный менеджер
        print("\n📋 Способ 1: Контекстный менеджер")
        with SimpleExcelAnalyzer(file_path) as analyzer:
            analyzer.print_summary()

        # Способ 2: Быстрая функция
        print("\n⚡ Способ 2: Быстрая функция")
        info = quick_analyze(file_path)
        if 'error' in info:
            print(f"❌ Ошибка: {info['error']}")
        else:
            print(f"✅ Файл: {info['file_name']}")
            print(f"📊 Листов: {info['sheets_count']}")
            print(f"📋 Названия: {info['sheet_names']}")

        # Способ 3: Прямое использование
        print("\n📖 Способ 3: Прямое использование")
        analyzer = SimpleExcelAnalyzer(file_path)
        basic_info = analyzer.get_basic_info()
        print(f"Основная информация: {basic_info}")
        analyzer.close()

    except FileNotFoundError:
        print("❌ Файл не найден")
        print("\n💡 Для тестирования:")
        print("1. Создайте Excel файл (.xlsx)")
        print("2. Откройте его в Excel")
        print("3. Запустите скрипт")
        print("4. read_only=True должен позволить чтение")

    except Exception as e:
        print(f"❌ Ошибка: {e}")

    print(f"\n🎯 ВЫВОДЫ:")
    print(f"✅ read_only=True решает большинство проблем")
    print(f"✅ Простой код = меньше багов")
    print(f"✅ openpyxl достаточно умный сам по себе")
    print(f"⚠️ Сложная логика нужна только для edge cases")
