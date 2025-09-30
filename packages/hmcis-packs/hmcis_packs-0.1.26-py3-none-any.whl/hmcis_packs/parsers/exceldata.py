import hashlib
import warnings
from collections import namedtuple
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import openpyxl
import pandas as pd
from openpyxl import load_workbook

from hmcis_packs.clean.cleaner import DataframeCleaner
from hmcis_packs.logger.logger_config import setup_logger

logger = setup_logger(__name__)


@dataclass
class WorkbookStructure:
    """Структура рабочей книги для сравнения."""

    sheet_names: Tuple[str, ...]  # Имена листов
    sheet_count: int  # Количество листов
    file_extension: str  # Расширение файла
    structure_hash: str = ""  # Хеш структуры для быстрого сравнения

    def __post_init__(self):
        # Генерируем хеш структуры для быстрого сравнения
        structure_data = f"{self.sheet_count}:{':'.join(sorted(self.sheet_names))}:{self.file_extension}"
        self.structure_hash = hashlib.md5(structure_data.encode()).hexdigest()[:8]

    def __eq__(self, other) -> bool:
        if not isinstance(other, WorkbookStructure):
            return False
        return self.structure_hash == other.structure_hash

    def __hash__(self) -> int:
        return hash(self.structure_hash)

    def is_similar_to(
        self,
        other: "WorkbookStructure",
        exact_match: bool = True,
        ignore_sheet_order: bool = True,
    ) -> bool:
        """
        Проверяет схожесть структур.

        Args:
            other: Другая структура для сравнения
            exact_match: Точное совпадение имен листов
            ignore_sheet_order: Игнорировать порядок листов
        """
        if not isinstance(other, WorkbookStructure):
            return False

        # Быстрая проверка по хешу для точного совпадения
        if exact_match and ignore_sheet_order:
            return self == other

        # Проверка количества листов
        if self.sheet_count != other.sheet_count:
            return False

        # Проверка имен листов
        if exact_match:
            if ignore_sheet_order:
                return set(self.sheet_names) == set(other.sheet_names)
            else:
                return self.sheet_names == other.sheet_names

        # Частичное совпадение (можно расширить логику)
        return len(set(self.sheet_names) & set(other.sheet_names)) > 0


SheetInfo = namedtuple(
    "SheetInfo", ["name", "index", "used_range", "row_count", "col_count"]
)


class ExcelParser:
    """
    Reads an Excel sheet using openpyxl, cleans it, and returns a DataFrame.
    Cross-platform alternative to COM-based Excel reading.

    Parameters:
      filepath: path to the .xlsx file
      sheet: worksheet to read (0-based index or substring of name)
      index_value: row-label to use as index
      retain_duplicates: if False, will drop duplicate columns
      read_only: if True, opens file in read-only mode (default True)
    """

    def __init__(
        self,
        filepath: Union[str, Path],
        sheet: Union[str, int],
        index_value: str,
        *,
        retain_duplicates: bool = False,
        read_only: bool = True,
    ) -> None:
        self.filepath = Path(filepath)
        self.sheet = sheet
        self.index_value = index_value
        self.retain_duplicates = retain_duplicates
        self.read_only = read_only
        self._wb = None
        self._structure: Optional[WorkbookStructure] = None

    def __eq__(self, other) -> bool:
        """Сравнение экземпляров по структуре книги."""
        if not isinstance(other, ExcelParser):
            return False

        # Если структуры еще не определены, определяем их
        if self._structure is None:
            self.get_structure()
        if other._structure is None:
            other.get_structure()

        return self._structure == other._structure

    def __hash__(self) -> int:
        """Хеш для использования в множествах и словарях."""
        if self._structure is None:
            self.get_structure()
        return hash(self._structure)

    def __repr__(self) -> str:
        return f"ExcelParser('{self.filepath.name}', sheet={self.sheet})"

    def __enter__(self):
        """Вход в контекстный менеджер - открывает Excel файл."""
        logger.info("Opening Excel file: %s", self.filepath)

        try:
            # Подавляем предупреждения openpyxl о стилях и других элементах
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore", category=UserWarning, module="openpyxl"
                )

                self._wb = load_workbook(
                    filename=str(self.filepath),
                    read_only=self.read_only,
                    data_only=True,  # Читаем только значения, не формулы
                )

            logger.info(
                "Sheets count: %s, names: %s",
                len(self._wb.sheetnames),
                self._wb.sheetnames,
            )
            return self

        except Exception as e:
            logger.exception("Failed to open Excel file: %s", e)
            self._cleanup()
            raise

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Выход из контекстного менеджера - закрывает рабочую книгу."""
        logger.info("Closing Excel file")
        self._cleanup()

        if exc_type is not None:
            logger.error("Exception occurred in ExcelParser context: %s", exc_val)

        return False  # Не подавляем исключения

    def _cleanup(self):
        """Очистка ресурсов."""
        if self._wb is not None:
            try:
                self._wb.close()
                logger.info("Closed workbook")
            except Exception as e:
                logger.warning("Error closing workbook: %s", e)
            finally:
                self._wb = None

    def _resolve_sheet(self, wb=None):
        """
        Возвращает worksheet объект по индексу или по подстроке в имени.
        Примечание: используем 0-based индексы в отличие от COM версии.
        """
        wb = wb or self._wb
        if wb is None:
            raise RuntimeError("Workbook is not open. Use within context manager.")

        if isinstance(self.sheet, int):
            try:
                if self.sheet < 0 or self.sheet >= len(wb.sheetnames):
                    raise IndexError(
                        f"Sheet index {self.sheet} is out of range (0-{len(wb.sheetnames) - 1})"
                    )

                sheet_name = wb.sheetnames[self.sheet]
                sheet = wb[sheet_name]
                logger.info(
                    "Resolved sheet by 0-based index %d: %s", self.sheet, sheet.title
                )
                return sheet

            except Exception as e:
                raise ValueError(f"Error accessing sheet at index {self.sheet}: {e}")
        else:
            name_sub = str(self.sheet)
            matches = [name for name in wb.sheetnames if name_sub in name]

            if not matches:
                raise ValueError(
                    f"No sheet name containing '{name_sub}' found. "
                    f"Available sheets: {wb.sheetnames}"
                )

            if len(matches) > 1:
                logger.warning(
                    "Multiple sheets matched '%s': %s; using the first: %s",
                    name_sub,
                    matches,
                    matches[0],
                )

            sheet = wb[matches[0]]
            logger.info("Resolved sheet by name: %s", sheet.title)
            return sheet

    def get_sheet(self):
        """Получить текущий лист (должен использоваться в контексте)."""
        if self._wb is None:
            raise RuntimeError("Workbook is not open. Use within context manager.")
        return self._resolve_sheet()

    def get_workbook(self):
        """Получить текущую рабочую книгу (должен использоваться в контексте)."""
        if self._wb is None:
            raise RuntimeError("Workbook is not open. Use within context manager.")
        return self._wb

    def get_structure(self, force_reload: bool = False) -> WorkbookStructure:
        """
        Получает структуру рабочей книги.

        Args:
            force_reload: Принудительно перезагрузить структуру
        """
        if self._structure is not None and not force_reload:
            return self._structure

        # Если книга уже открыта, используем ее
        if self._wb is not None:
            self._structure = self._extract_structure_from_wb(self._wb)
        else:
            # Быстро открываем книгу только для получения структуры
            with self as parser:
                self._structure = self._extract_structure_from_wb(parser._wb)

        logger.info(
            "Extracted structure: %s sheets, hash=%s",
            self._structure.sheet_count,
            self._structure.structure_hash,
        )
        return self._structure

    def _extract_structure_from_wb(self, wb) -> WorkbookStructure:
        """Извлекает структуру из открытой рабочей книги."""
        sheet_names = tuple(wb.sheetnames)
        sheet_count = len(wb.sheetnames)
        file_extension = self.filepath.suffix.lower()

        return WorkbookStructure(
            sheet_names=sheet_names,
            sheet_count=sheet_count,
            file_extension=file_extension,
        )

    def get_detailed_sheet_info(self) -> List[SheetInfo]:
        """Получает подробную информацию о каждом листе."""
        if self._wb is None:
            with self as parser:
                return parser._get_detailed_info_from_wb()
        else:
            return self._get_detailed_info_from_wb()

    def _get_detailed_info_from_wb(self) -> List[SheetInfo]:
        """Извлекает подробную информацию о листах."""
        sheets_info = []

        for i, sheet_name in enumerate(self._wb.sheetnames):
            try:
                sheet = self._wb[sheet_name]

                # Определяем используемый диапазон
                if sheet.max_row == 1 and sheet.max_column == 1:
                    # Проверяем, есть ли данные в единственной ячейке
                    if sheet.cell(1, 1).value is None:
                        row_count = col_count = 0
                        range_address = "Empty"
                    else:
                        row_count = col_count = 1
                        range_address = "A1"
                else:
                    row_count = sheet.max_row
                    col_count = sheet.max_column
                    range_address = (
                        f"A1:{openpyxl.utils.get_column_letter(col_count)}{row_count}"
                    )

                sheet_info = SheetInfo(
                    name=sheet_name,
                    index=i,  # 0-based в отличие от COM версии
                    used_range=range_address,
                    row_count=row_count,
                    col_count=col_count,
                )
                sheets_info.append(sheet_info)

            except Exception as e:
                logger.warning("Error reading sheet %s: %s", sheet_name, e)

        return sheets_info

    def is_similar_to(self, other: "ExcelParser", **kwargs) -> bool:
        """
        Проверяет схожесть структур с другим экземпляром.

        Args:
            other: Другой экземпляр ExcelParser
            **kwargs: Параметры для is_similar_to метода WorkbookStructure
        """
        if not isinstance(other, ExcelParser):
            return False

        my_structure = self.get_structure()
        other_structure = other.get_structure()

        return my_structure.is_similar_to(other_structure, **kwargs)

    def compare_with(self, other: "ExcelParser") -> Dict[str, any]:
        """
        Подробное сравнение с другим экземпляром.

        Returns:
            Словарь с результатами сравнения
        """
        if not isinstance(other, ExcelParser):
            return {"error": "Can only compare with another ExcelParser"}

        my_structure = self.get_structure()
        other_structure = other.get_structure()

        comparison = {
            "files": {"self": self.filepath.name, "other": other.filepath.name},
            "identical_structure": my_structure == other_structure,
            "structure_hashes": {
                "self": my_structure.structure_hash,
                "other": other_structure.structure_hash,
            },
            "sheet_counts": {
                "self": my_structure.sheet_count,
                "other": other_structure.sheet_count,
                "match": my_structure.sheet_count == other_structure.sheet_count,
            },
            "sheet_names": {
                "self": my_structure.sheet_names,
                "other": other_structure.sheet_names,
                "common": tuple(
                    set(my_structure.sheet_names) & set(other_structure.sheet_names)
                ),
                "unique_to_self": tuple(
                    set(my_structure.sheet_names) - set(other_structure.sheet_names)
                ),
                "unique_to_other": tuple(
                    set(other_structure.sheet_names) - set(my_structure.sheet_names)
                ),
            },
            "extensions": {
                "self": my_structure.file_extension,
                "other": other_structure.file_extension,
                "match": my_structure.file_extension == other_structure.file_extension,
            },
        }

        return comparison

    def read_data(self) -> pd.DataFrame:
        """Основной публичный метод: открыть Excel, выбрать лист, вернуть очищенный DataFrame."""
        # Если уже в контексте, используем открытую книгу
        if self._wb is not None:
            return self._read_data_from_open_wb()

        # Иначе открываем временно через контекстный менеджер
        with self as parser:
            return parser._read_data_from_open_wb()

    def _read_data_from_open_wb(self) -> pd.DataFrame:
        """Читает данные из уже открытой рабочей книги."""
        sheet = self._resolve_sheet()
        logger.info("Processing data on sheet -> %s", sheet.title)
        return self._process_sheet_data(sheet)

    def _process_sheet_data(self, sheet) -> pd.DataFrame:
        """Обрабатывает данные с листа."""
        # Получаем все данные из листа
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        if not data:
            logger.warning("Sheet is empty")
            return pd.DataFrame()

        df = pd.DataFrame(data)
        logger.info("Raw data shape: %s", df.shape)

        # Очищаем данные
        cleaner = DataframeCleaner(df)
        cleaner.adj_by_row_index(self.index_value)

        if not self.retain_duplicates:
            cleaner.remove_duplicated_cols()
            logger.info("Removed duplicated columns")

        logger.info("Final cleaned data shape: %s", cleaner.df.shape)
        return cleaner.df

    def read_data_pandas(self, **pandas_kwargs) -> pd.DataFrame:
        """
        Альтернативный метод чтения через pandas.read_excel().
        Может быть полезен для простых случаев.

        Args:
            **pandas_kwargs: Дополнительные параметры для pandas.read_excel()
        """
        logger.info("Reading data using pandas.read_excel()")

        # Определяем имя листа для pandas
        sheet_name = None
        if isinstance(self.sheet, int):
            # pandas использует 0-based индексы, как и наша реализация
            sheet_name = self.sheet
        else:
            # Для строкового идентификатора нужно найти точное имя
            with self as parser:
                resolved_sheet = parser._resolve_sheet()
                sheet_name = resolved_sheet.title

        # Стандартные параметры
        default_kwargs = {"sheet_name": sheet_name, "engine": "openpyxl"}
        default_kwargs.update(pandas_kwargs)

        try:
            df = pd.read_excel(str(self.filepath), **default_kwargs)
            logger.info("Data loaded via pandas: %s", df.shape)

            # Применяем очистку, если нужно
            if hasattr(self, "index_value") and self.index_value:
                cleaner = DataframeCleaner(df)
                cleaner.adj_by_row_index(self.index_value)

                if not self.retain_duplicates:
                    cleaner.remove_duplicated_cols()
                    logger.info("Removed duplicated columns")

                return cleaner.df

            return df

        except Exception as e:
            logger.error("Failed to read data via pandas: %s", e)
            raise


if __name__ == "__main__":
    # Пример 1: Традиционное использование (обратите внимание: индексы теперь 0-based)
    print("=== Традиционное использование ===")
    parser = ExcelParser(
        r"V:\Accounting\Work\Мерц\2025\3 квартал 2025\Август 2025 года\Отчетность\!Начисление МСФО_август 2025.xlsx",
        # Замените на реальный путь
        sheet=0,  # или используйте 0-based индекс, например: sheet=0
        index_value="a1",
        retain_duplicates=False,
    )
    # df = parser.read_data()
    # print(f"Результат: {df.shape}")

    # Пример 2: Использование как контекстный менеджер
    print("\n=== Использование как контекстный менеджер ===")
    try:
        with (
            ExcelParser(
                r"V:\Accounting\Work\Мерц\2025\3 квартал 2025\Август 2025 года\Отчетность\!Начисление МСФО_август 2025.xlsx",  # Замените на реальный путь
                sheet=0,  # 0-based индекс (первый лист)
                index_value="Отдел инициатор",
            ) as parser
        ):
            # Получаем информацию о структуре
            structure = parser.get_structure()
            print(f"Структура: {structure.sheet_count} листов")

            # Получаем подробную информацию о листах
            sheets_info = parser.get_detailed_sheet_info()
            for info in sheets_info:
                print(
                    f"Лист: {info.name}, строк: {info.row_count}, столбцов: {info.col_count}"
                )

            # Читаем данные
            # df = parser.read_data()
            # print(f"Результат через контекстный менеджер: {df.shape}")

    except FileNotFoundError:
        print("Файл не найден - замените путь на реальный файл для тестирования")

    # Пример 3: Использование pandas метода
    print("\n=== Использование метода pandas ===")
    try:
        parser_pandas = ExcelParser(
            r"V:\Accounting\Work\Мерц\2025\3 квартал 2025\Август 2025 года\Отчетность\!Начисление МСФО_август 2025.xlsx",
            sheet=0,
            index_value="",
            retain_duplicates=True,
        )
        # df_pandas = parser_pandas.read_data_pandas(header=0)
        # print(f"Результат через pandas: {df_pandas.shape}")
    except FileNotFoundError:
        print("Файл не найден - замените путь на реальный файл для тестирования")

    print("=== Тест завершен ===")
